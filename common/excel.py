#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Date       : 2019-08-05 14:15:28
# @Author     : caryangBingo
# @Filename   : excel.py
# @Version    : Version 1.0

import io
import os
import sys
import json
import openpyxl
import itertools
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from common import consts
from common.OAT import *
# sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


excel_filename = os.path.dirname(os.path.dirname(
    os.path.abspath(__file__))) + '/data/broker-testcase_src.xlsx'


class OrthogonalCaseProcessing(object):
    """docstring for OrthogonalCaseProcessing"""

    def __init__(self):
        self.oat = OAT()
        self.workbook = openpyxl.load_workbook(filename=excel_filename)
        self.worksheets = self.workbook.worksheets
        self.rows_cut = consts.ROWS_CUT
        self.cols_cut = consts.COLS_CUT
        self.case_sig = consts.CASE_SIG
        self.priority = consts.API_ATTR.get('priority')
        self.tags = consts.API_ATTR.get('tags')
        self.apipath = consts.API_ATTR.get('apipath')
        self.method = consts.API_ATTR.get('method')
        self.summary = consts.API_ATTR.get('summary')
        self.args = consts.API_ATTR.get('args')

    def getWorksheets(self):
        self.worksheet = self.worksheets[0]
        return self.worksheet

    def readAttributes(self, worksheet):
        self.attrDict = {}
        self.attrDict['priority'] = worksheet[self.priority].value
        self.attrDict['tags'] = worksheet[self.tags].value
        self.attrDict['apipath'] = worksheet[self.apipath].value
        self.attrDict['method'] = worksheet[self.method].value
        self.attrDict['summary'] = worksheet[self.summary].value
        self.attrDict['args'] = worksheet[self.args].value
        return self.attrDict

    def readWorkslData(self, worksheet):

        self.sheet_rows = worksheet.max_row  # 获取行数
        self.sheet_cols = worksheet.max_column  # 获取列数

        self.req_paramsRows = list(
            map(lambda x: str(x + self.rows_cut + 1), range(self.sheet_rows - self.rows_cut)))

        # 参数值行
        self.req_paramsCols = list(chr(i + 65)
                                   for i in range(self.case_sig, self.sheet_cols - self.cols_cut))
        # 期望值行
        self.req_expectCols = list(chr(i + 65)
                                   for i in range(self.sheet_cols - self.cols_cut, self.sheet_cols))
        self.req_expectCols.insert(0, chr(self.case_sig - 1 + 65))

        # 参数值名
        self.req_paramsName = list(
            map(lambda x: x + str(self.rows_cut), self.req_paramsCols))
        # 期望值名
        self.req_expectName = list(
            map(lambda x: x + str(self.rows_cut), self.req_expectCols))

        # 参数值map
        self.params_map = [(key, list(map(lambda x:''.join(x), zip(
            [key] * len(self.req_paramsRows), self.req_paramsRows)))) for key in self.req_paramsCols]

        # 期望值map
        self.expect_map = [(key, list(map(lambda x:''.join(x), zip(
            [key] * len(self.req_paramsRows), self.req_paramsRows)))) for key in self.req_expectCols]

        # 获取正交表用例cell
        self.case_cell = json.loads(json.dumps(
            self.oat.genSets(OrderedDict(self.params_map))))

        # 填充参数值用例数据
        self.params_cellData = list(list(map(lambda x: celldata.get(x), self.req_paramsCols))
                                    for celldata in self.case_cell)
        self.params_cellData = list(
            list(map(lambda x: str(worksheet[x].value), data)) for data in self.params_cellData)

        # 填充期望值用例数据
        self.expect_cellData = list(
            list(map(lambda x: str(x), data[1])) for data in self.expect_map)
        self.expect_cellData = list(list(map(lambda x: worksheet[x[i]].value, self.expect_cellData)) for i in range(
            len(self.req_paramsRows)))

        # 参数值名
        self.req_paramsName = list(
            map(lambda x: worksheet[x].value, self.req_paramsName))
        # 期望值名
        self.req_expectName = list(
            map(lambda x: worksheet[x].value, self.req_expectName))

        # self.req_paramsData = list(map(lambda x: dict(
        # zip(self.req_paramsName, x[0])),
        # itertools.groupby(self.params_cellData)))
        self.req_paramsData = list(map(lambda x: dict(
            zip(self.req_paramsName, x)), self.params_cellData))

        self.req_expectData = list(map(lambda x: dict(
            zip(self.req_expectName, x)), self.expect_cellData))

        # 计算期望值和参数值匹配的数量
        self.req_expectNums = int(len(
            self.req_paramsData)) // int(len(self.req_expectData))

        self.req_expectData = self.req_expectData * self.req_expectNums + self.req_expectData[:int(len(
            self.req_paramsData) - len(self.req_expectData * self.req_expectNums))]

        self.req_paramsExpe = list(
            map(lambda x: x, zip(self.req_paramsData, self.req_expectData)))

        # print(self.req_paramsExpe)
        # print(self.req_paramsData)
        # return self.req_paramsExpe


if __name__ == '__main__':
    CaseProcessing = OrthogonalCaseProcessing()
    CaseProcessing.readWorkslData(CaseProcessing.getWorksheets())
    # print(CaseProcessing.readAttributes(CaseProcessing.getWorksheets()))
