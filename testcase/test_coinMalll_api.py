#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Date       : 2019-07-30 09:54:56
# @Author     : caryangBingo
# @Filename   : example.py
# @Version    : Version 1.0

import io
import os
import sys
import json
import requests
import pysnooper
import allure
import pytest
import yaml
import openpyxl
import itertools
sys.path.append("../common/")
from assertpy.assertpy import assert_that, fail
# from ddt import ddt, file_data
from OrthogonalArray.OAT import *
# sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
#
oat = OAT()
coinMall_url = "http://218.244.157.23:8080/coinMall"
apiPath = "/api/getCommentList"
headers = {
    "Accept":
    "application/json;charset=UTF-8",
    'User-Agent':
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0'
}

excel_filename = '../data/broker-testcase_src.xlsx'

# with open("../data/getCommentListUsingGET.yaml", 'r', encoding='utf-8') as fp:
#    yaml_data = yaml.safe_load(fp)

# print(yaml_data)

# http://devmsmobapi.lanmang.me/api/EnterpriseQueryMobile/GetEnterpriseById?entId=1042&uid=20165943


class OrthogonalCaseProcessing(object):
    """docstring for OrthogonalCaseProcessing"""

    def __init__(self):
        self.workbook = openpyxl.load_workbook(filename=excel_filename)

    def readExcelData(self):
        # case2 = OrderedDict([('A', ['A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9']),
        #                     ('B', ['B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B9']),
        #                     ('C', ['C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9']),
        #                     ('D', ['D3', 'D4', 'D5', 'D6', 'D7', 'D8', 'D9']),
        #                     ('E', ['E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'E9'])])
        # exp_case2 = OrderedDict([('F', ['F3', 'F4']), ('G', ['G3', 'G4'])])

        # sheet_cell = json.loads(json.dumps(oat.genSets(case2)))
        # sheet_exp = json.loads(json.dumps(oat.genSets(exp_case2)))

        sheet = self.workbook.active
        rows = sheet.max_row  # 获取行数
        cols = sheet.max_column  # 获取列数

        req_paramsRows = list(map(str, (range(3, rows + 1))))
        req_paramsCols = list(chr(i + 65) for i in range(0, cols - 2))
        req_paramsName = list(map(lambda x: x + str(2), req_paramsCols))

        case_map = [(key, list(map(lambda x:''.join(x), zip(
            [key] * len(req_paramsRows), req_paramsRows)))) for key in req_paramsCols]

        case2 = OrderedDict(case_map)
        sheet_cell = json.loads(json.dumps(oat.genSets(case2)))

        """
        for col in req_paramsCols:
            col_list = []
            for row in req_paramsRows:
                col_list.append(col + row)
            yield ("('{}': {})".format(col, col_list))
            # print('{}: {}'.format(col, col + str(row)))
        """
        cellData = list(list(map(lambda x: cellData.get(x), req_paramsCols))
                        for cellData in sheet_cell)

        cellData = list(
            list(map(lambda x: str(sheet[x].value), data)) for data in cellData)

        req_paramsName = list(map(lambda x: sheet[x].value, req_paramsName))

        # params_case2 += '{'
        # params_case2 += '"aboutId": {0}, "uid": {1}, "parentId": {2}, "pageIndex": {3}, "pageSize": {4}'.format(
        #    sheet[A_cellDict].value, sheet[B_cellDict].value, sheet[C_cellDict].value, sheet[D_cellDict].value, sheet[E_cellDict].value)
        # params_case2 += '}'
        req_paramsData = list(map(lambda x: dict(
            zip(req_paramsName, x[0])), itertools.groupby(cellData)))

        yield req_paramsData

    # readExcelData()
    #print([i for i in readExcelData()])

    """
    for cellDict in sheet_exp:
        F_cellDict = cellDict.get('F')
        G_cellDict = cellDict.get('G')
        print(C_cellDict, D_cellDict)
        print(sheet[C_cellDict].value, sheet[D_cellDict].value)
    """

    @allure.step
    @pytest.mark.parametrize('params', [i for i in CaseProcessing.readExcelData(self)][0])
    def test_api_getCommentList_request(self, params):
        print(params)
        r = requests.get(coinMall_url + apiPath,
                         params=params,
                         headers=headers)
        reqUrl = r.url
        result = r.json()
        # elapsedTime = r.elapsed.total_seconds()
        # assertEqual(r.status_code, http_statusCode)
        # assertEqual(r.status_code, http_statusCode)
        # assert result['code'] == 0
        try:
            assert_that(result.get('code')).is_equal_to(0)
        except AssertionError:
            assert_that(result.get('code')).is_equal_to(-1)
        # assert Assertions.assert_msg(
        #    result['message'], responsesBody['message'])
        # assert Assertions.assert_body(
        #    result['bodyMessage'], responsesBody['bodyMessage'])
        # assert Assertions.assert_subCode(
        #    result['subCode'], responsesBody['subCode']


if __name__ == '__main__':
    CaseProcessing = OrthogonalCaseProcessing()
